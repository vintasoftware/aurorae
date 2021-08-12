from openpyxl import load_workbook


def worksheet_dict_reader(worksheet):
    """
    A generator for the rows in a given worksheet. It maps columns on
    the first row of the spreadsheet to each of the following lines
    returning a dict like {
        "header_column_name_1": "value_column_1",
        "header_column_name_1": "value_column_2"
    }
    :param worksheet: a worksheet object
    :type worksheet: `openpyxl.worksheet.ReadOnlyWorksheet`
    """
    rows = worksheet.iter_rows(values_only=True)

    # pylint: disable=stop-iteration-return
    header = list(filter(None, next(rows)))
    for row in rows:
        if not any(row):
            return
        yield dict(zip(header, row))


def get_spreadsheet_data(filename) -> dict:
    """
    Uses `openpyxl.load_workbook` to process the specified file.
    Returns a dict of the spreadsheet data grouped by worksheet.
    """
    workbook = load_workbook(filename=filename, read_only=True, data_only=True)
    dados_empresa = worksheet_dict_reader(workbook["Empresa"])
    dados_funcionarios = worksheet_dict_reader(workbook["Funcionários"])
    dados_pagamentos = worksheet_dict_reader(workbook["Pagamentos"])

    return {
        "Empresa": list(dados_empresa)[0],
        "Funcionários": list(dados_funcionarios),
        "Pagamentos": list(dados_pagamentos),
    }
