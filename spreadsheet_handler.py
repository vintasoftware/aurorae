from openpyxl import load_workbook

from cnab.cnab240.v10_7 import lambdas
from cnab.cnab240.v10_7.spreadsheet_map import CUSTOM_FIELDS_MAPPING, MODELS_SPREADSHEET_MAP


INITIAL_DATA_DICT = {
    "header": [],
    "trailer": [],
    "lote_header": [],
    "lote_trailer": [],
    "lote_detalhe_segmento_c": [],
    "lote_detalhe_segmento_b": [],
    "lote_detalhe_segmento_a": [],
}


def worksheet_dict_reader(worksheet):
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)
    for row in rows:
        if not any(row):
            return
        yield dict(zip(header, row))


def get_spreadsheet_data():
    workbook = load_workbook(
        filename="./tmp/test_data.xlsx", read_only=True, data_only=True
    )
    dados_empresa = worksheet_dict_reader(workbook["Empresa"])
    dados_funcionarios = worksheet_dict_reader(workbook["Funcionários"])
    dados_pagamentos = worksheet_dict_reader(workbook["Pagamentos"])

    return {
        "Empresa": list(dados_empresa),
        "Funcionários": list(dados_funcionarios),
        "Pagamentos": list(dados_pagamentos),
    }


def get_initial_data_from(spreadsheet_data):
    initial_data = INITIAL_DATA_DICT.copy()
    invalid_field_maps = []

    # Models
    for segment_name, fields in MODELS_SPREADSHEET_MAP.items():
        model_initial_data = {}
        # Fields
        for field_name, field_specs in fields.items():
            # Sheet Rows
            registers = []
            sheet_rows = spreadsheet_data[field_specs["sheet_name"]]
            for row in sheet_rows:
                try:
                    if isinstance(field_specs["column_name"], list):
                        # Multiple fields mapped as one column
                        custom_fields = {}
                        for mapped_column in field_specs["column_name"]:
                            custom_column_name = mapped_column[0]
                            try:
                                custom_fields[custom_column_name] = row[
                                    custom_column_name
                                ]
                            except KeyError:
                                error_msg = (
                                    f"The column '{custom_column_name}' doesn't "
                                    f"exists on the '{field_specs['sheet_name']}' sheet."
                                )
                                invalid_field_maps.append({field_name: error_msg})
                        model_initial_data[field_name] = custom_fields
                    else:
                        data = row[field_specs["column_name"]]
                        model_initial_data[field_name] = str(data)

                    if "lote" in segment_name:
                        registers += [model_initial_data]
                except KeyError:
                    error_msg = (
                        f"The column '{field_specs['column_name']}' doesn't "
                        f"exists on the '{field_specs['sheet_name']}' sheet."
                    )
                    invalid_field_maps.append({field_name: error_msg})

        if registers:
            initial_data[segment_name] = registers
        else:
            initial_data[segment_name] = [model_initial_data]

    if invalid_field_maps:
        raise Exception(invalid_field_maps)

    return initial_data


def _generate_line(fields, spreadsheet_data):
    invalid_field_maps = []
    model_initial_data = {}
    for field_name, field_specs in fields.items():
        lambda_func_name = field_specs["lambda"]
        try:
            if lambda_func_name == "default":
                # For the fields to get the default value this values must be none or ''
                # doing this to ease development
                model_initial_data[field_name] = ""
            else:
                has_params = field_specs.get("params", False)
                method_to_call = getattr(lambdas, lambda_func_name)

                if has_params:
                    model_initial_data[field_name] = method_to_call(spreadsheet_data)
                else:
                    model_initial_data[field_name] = method_to_call()
        except KeyError:
            error_msg = (
                f"The column '{field_specs['lambda']}' doesn't "
                f"exists on the '{lambdas.__file__}' sheet."
            )
            invalid_field_maps.append({field_name: error_msg})
    return model_initial_data, invalid_field_maps


def get_calculated_fields_data(spreadsheet_data):
    """
    Fields need to be generated in sequence. In particular, the trailers and the lote_detalhe_segmentos
    need information from previous lines.
    """
    initial_data = INITIAL_DATA_DICT.copy()
    invalid_field_maps = []

    for segment_name in ["header", "lote_header"]:
        fields = CUSTOM_FIELDS_MAPPING[segment_name]
        model_initial_data, errs = _generate_line(fields, spreadsheet_data)
        initial_data[segment_name] = [model_initial_data]
        invalid_field_maps += errs

    for i, _ in enumerate(spreadsheet_data["lote_detalhe_segmento_a"]):
        for segment_name in [
            "lote_detalhe_segmento_a",
            "lote_detalhe_segmento_b",
            "lote_detalhe_segmento_c",
        ]:
            fields = CUSTOM_FIELDS_MAPPING[segment_name]
            model_initial_data, errs = _generate_line(fields, spreadsheet_data)
            invalid_field_maps += errs
            initial_data[segment_name] += [model_initial_data]

    for segment_name in ["lote_trailer", "trailer"]:
        fields = CUSTOM_FIELDS_MAPPING[segment_name]
        model_initial_data, errs = _generate_line(fields, spreadsheet_data)
        initial_data[segment_name] = [model_initial_data]
        invalid_field_maps += errs

    if invalid_field_maps:
        raise Exception(invalid_field_maps)

    return initial_data


def generate_initial_data():
    keys = INITIAL_DATA_DICT.keys()
    initial_data = INITIAL_DATA_DICT.copy()
    spreadsheet_data = get_spreadsheet_data()

    input_fields_data = get_initial_data_from(spreadsheet_data)
    custom_fields_data = get_calculated_fields_data(input_fields_data)

    for key in keys:
        input_fields_content = input_fields_data[key]
        custom_fields_content = custom_fields_data[key]

        assert len(input_fields_content) == len(custom_fields_content)

        for _input, _custom in zip(input_fields_content, custom_fields_content):
            segment = {}
            segment.update(_input)
            segment.update(_custom)
            initial_data[key].append(segment)

    return initial_data


def generate_initial_data_with_connectors():
    from connectors.worksheet_handler import parse_data_from

    keys = INITIAL_DATA_DICT.keys()
    initial_data = INITIAL_DATA_DICT.copy()
    spreadsheet_data = get_spreadsheet_data()

    input_fields_data = parse_data_from(spreadsheet_data)
    custom_fields_data = get_calculated_fields_data(input_fields_data)

    for key in keys:
        input_fields_content = input_fields_data[key]
        input_fields_content = [info for info in input_fields_content if info]
        custom_fields_content = custom_fields_data[key]

        if key == "lote_header" or key == "lote_trailer":
            input_fields_content = [input_fields_content[0]]

        assert len(input_fields_content) == len(
            custom_fields_content
        ), f"{key}: has different count of input fields and custom fields"

        for _input, _custom in zip(input_fields_content, custom_fields_content):
            segment = {}
            segment.update(_input)
            segment.update(_custom)
            initial_data[key].append(segment)

    return initial_data
