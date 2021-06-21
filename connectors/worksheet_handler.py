from pathlib import Path
from openpyxl import load_workbook


def worksheet_dict_reader(worksheet):
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)
    for row in rows:
        if not any(row):
            return
        yield dict(zip(header, row))


def get_spreadsheet_data(filename: Path) -> dict:
    workbook = load_workbook(filename=filename, read_only=True, data_only=True)
    dados_empresa = worksheet_dict_reader(workbook["Empresa"])
    dados_funcionarios = worksheet_dict_reader(workbook["Funcionários"])
    dados_pagamentos = worksheet_dict_reader(workbook["Pagamentos"])

    return {
        "Empresa": list(dados_empresa),
        "Funcionários": list(dados_funcionarios),
        "Pagamentos": list(dados_pagamentos),
    }


def parse_data_from(spreadsheet_data: dict, spreadsheet_map: dict) -> dict:
    from connectors.parser import fill_segment_data, get_field_based_on
    errors = []
    parsed_data = {}
    amount_of_payments = len(spreadsheet_data["Pagamentos"])

    for segment_name, segment_fields in spreadsheet_map.items():
        for field_name, field_specs in segment_fields.items():
            sheet_name = field_specs["sheet_name"]
            related_column_name = field_specs["column_name"]
            sheet_rows = spreadsheet_data[sheet_name]

            data, invalid_field_maps = get_field_based_on(
                field_name=field_name,
                origin_spreadsheet_name=related_column_name,
                sheet_rows=sheet_rows,
                amount_of_payments=amount_of_payments,
            )

            if invalid_field_maps:
                errors += invalid_field_maps
                continue

            parsed_data = fill_segment_data(
                data=parsed_data,
                segment_name=segment_name,
                segment_value=data,
                amount_of_payments=amount_of_payments,
            )

    if invalid_field_maps:
        raise Exception(invalid_field_maps)

    return parsed_data

