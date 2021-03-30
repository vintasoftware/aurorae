from openpyxl import load_workbook

from cnab240.v10_7 import lambdas
from cnab240.v10_7.spreadsheet_map import MODELS_SPREADSHEET_MAP, CUSTOM_FIELDS_MAPPING


def worksheet_dict_reader(worksheet):
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)
    for row in rows:
        if not any(row):
            return

        yield dict(zip(header, row))


def get_spreadsheet_data():
    workbook = load_workbook(filename="./dados-pagamentos.xlsx", read_only=True, data_only=True)
    dados_empresa = worksheet_dict_reader(workbook["Empresa"])
    dados_funcionarios = worksheet_dict_reader(workbook["Funcionários"])
    dados_pagamentos = worksheet_dict_reader(workbook["Pagamentos"])

    return {
        "Empresa": list(dados_empresa),
        "Funcionários": list(dados_funcionarios),
        "Pagamentos": list(dados_pagamentos)
    }


def get_initial_data(spreadsheet_data):
    initial_data = {
        'lote_header': [],
        'lote_trailer': [], 
        'lote_detalhe_segmento_c': [], 
        'lote_detalhe_segmento_b': [],
        'lote_detalhe_segmento_a': [], 
    }
    invalid_field_maps = []

    # Models
    for segment_name, fields in MODELS_SPREADSHEET_MAP.items():
        model_initial_data = {}
        # Fields
        for field_name, field_specs in fields.items():
            # Sheet Rows
            registers = []
            sheet_rows = spreadsheet_data[field_specs["sheet_name"]]
            for row in sheet_rows:
                try:
                    if isinstance(field_specs["column_name"], list):
                        # Multiple fields mapped as one column
                        for mapped_column in field_specs["column_name"]:
                            custom_column_name = mapped_column[0]
                            try:
                                model_initial_data[field_name] = {
                                    custom_column_name: row[custom_column_name]
                                }
                            except KeyError:
                                error_msg = (
                                    f"The column '{custom_column_name}' doesn't "
                                    f"exists on the '{field_specs['sheet_name']}' sheet."
                                )
                                invalid_field_maps.append({field_name: error_msg})
                    else:
                        data = row[field_specs["column_name"]]
                        model_initial_data[field_name] = str(data)

                    if 'lote' in segment_name:
                        registers += [model_initial_data]
                except KeyError:
                    error_msg = (
                        f"The column '{field_specs['column_name']}' doesn't "
                        f"exists on the '{field_specs['sheet_name']}' sheet."
                    )
                    invalid_field_maps.append({field_name: error_msg})
        
        if registers:
            initial_data[segment_name] = registers
        else:
            initial_data[segment_name] = model_initial_data

    if invalid_field_maps:
        raise Exception(invalid_field_maps)

    return initial_data


def _generate_line(fields, spreadsheet_data):
    model_initial_data = {}
    for field_name, field_specs in fields.items():
        lambda_func_name = field_specs["lambda"]
        try:
            if lambda_func_name == 'default':
                # For the fields to get the default value this values must be none or ''
                # doing this to ease development
                model_initial_data[field_name] = 'default'
            else:
                has_params = field_specs.get("params", False)
                method_to_call = getattr(lambdas, lambda_func_name)

                if has_params:
                    model_initial_data[field_name] = method_to_call(spreadsheet_data)
                else:
                    model_initial_data[field_name] = method_to_call()
        except KeyError:
            error_msg = (
                f"The column '{field_specs['lambda']}' doesn't "
                f"exists on the '{lambdas.__file__}' sheet."
            )
            invalid_field_maps.append({field_name: error_msg})
    return model_initial_data


def get_custom_fields_data(initial_data, spreadsheet_data):
    initial_data = {
        'lote_header': [],
        'lote_trailer': [], 
        'lote_detalhe_segmento_c': [], 
        'lote_detalhe_segmento_b': [],
        'lote_detalhe_segmento_a': [], 
    }
    invalid_field_maps = []

    for segment_name, fields in CUSTOM_FIELDS_MAPPING.items():
        model_initial_data = {}

        if segment_name in ["header", "trailer", "lote_header", "lote_trailer"]:
            model_initial_data = _generate_line(fields, spreadsheet_data)
            initial_data[segment_name] = model_initial_data
            continue

        for i, _ in enumerate(spreadsheet_data['lote_detalhe_segmento_c']):
            model_initial_data = _generate_line(fields, spreadsheet_data)

            initial_data[segment_name] += [model_initial_data]

    if invalid_field_maps:
        raise Exception(invalid_field_maps)

    return initial_data


if __name__ == "__main__":
    spreadsheet_data = get_spreadsheet_data()
    fields_initial_data = get_initial_data(spreadsheet_data)
    # print(fields_initial_data)
    custom_fields_data = get_custom_fields_data(None, fields_initial_data)
    print(custom_fields_data)
